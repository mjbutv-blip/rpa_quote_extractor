import os
import openpyxl


BRA_MOLD_NOTE = (
    "价格中需要含一套正确模具费用，下单后，如果是由于客人原因更改模具，"
    "费用由我司支付，如果是由于贵厂的原因，导致重新开模具，费用由贵厂承担，请悉知"
)


def _to_plain_text(value) -> str:
    """写入单元格前的兜底拍平：防止上游意外返回 tuple/list/嵌套 dict 时把结构体写入 Excel。"""
    if value is None:
        return ""
    if isinstance(value, (tuple, list)):
        value = value[0] if value else ""
    if isinstance(value, dict):
        value = value.get("translated") or value.get("text") or value.get("value") or value.get("original") or ""
    return str(value).strip()


def _needs_bra_mold_note(product_name: str) -> bool:
    product_name = _to_plain_text(product_name)
    return any(keyword in product_name for keyword in ("文胸", "塞杯无缝内衣", "固定杯背心"))


def write_to_template(data_list, template_path, output_path):
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"找不到模板文件: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    target_sheet = next((s for s in wb.sheetnames if s.startswith("总表")), None)
    if target_sheet is None:
        raise ValueError(f"模板中未找到以'总表'开头的工作表，当前 Sheet 列表: {wb.sheetnames}")

    ws = wb[target_sheet]
    start_row = 5

    for data in data_list:
        ws.cell(row=start_row, column=2,  value=_to_plain_text(data.get("order_id", "")))
        ws.cell(row=start_row, column=6,  value=_to_plain_text(data.get("product_name", "")))
        ws.cell(row=start_row, column=10, value=_to_plain_text(data.get("fabric_quality", "")))
        ws.cell(row=start_row, column=11, value=_to_plain_text(data.get("color_print", "")))
        ws.cell(row=start_row, column=14, value=_to_plain_text(data.get("size_range", "")))
        if _needs_bra_mold_note(data.get("product_name", "")):
            ws.cell(row=start_row, column=17, value=BRA_MOLD_NOTE)

        start_row += 1

    wb.save(output_path)
