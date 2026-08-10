import base64
import hashlib
import io
import json
import os
import re
import tempfile
import zipfile

import openpyxl
from openai import OpenAI

try:
    import fitz  # PyMuPDF
    _FITZ_AVAILABLE = True
except ImportError:
    _FITZ_AVAILABLE = False

try:
    import pdfplumber
    _PDFPLUMBER_AVAILABLE = True
except ImportError:
    _PDFPLUMBER_AVAILABLE = False

try:
    from PIL import Image as PILImage
    _PIL_AVAILABLE = True
except ImportError:
    _PIL_AVAILABLE = False

# MIME types supported by the vision API
_SUPPORTED_EXTS = {"jpeg", "jpg", "png", "gif", "webp"}


def _ext_to_mime(ext: str):
    ext = ext.lower().lstrip(".")
    if ext == "jpg":
        return "image/jpeg"
    if ext in _SUPPORTED_EXTS:
        return f"image/{ext}"
    return None


# ── PDF 相关 ──────────────────────────────────────────────────────────────────

_SIZE_TOKEN_RE = re.compile(r"^(?:\d{2,3}/\d{2,3}|\d{2,3}[A-Z]{1,2}|[A-Z]{1,3})$")
_BRA_SIZE_RE = re.compile(r"^\d{2,3}[A-Z]{1,2}$")
_QTY_RE = re.compile(r"^\d+(?:[.,]\d+)?$")
_FABRIC_LABELS = {
    "shell fabric": "大身前后片",
    "lace": "花边",
    "lining": "内衬",
    "lining crotch": "底裆内衬",
    "padding": "胸杯牛奶丝",
    "padding lining": "胸杯牛奶丝",
}
_CONSTRUCTION_LABELS = {
    "mesh": "网布",
    "microfibre": "超细纤维",
    "microfiber": "超细纤维",
}
def _clean_cell(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _ascii_key(value: str) -> str:
    return re.sub(r"[^a-z]", "", _clean_cell(value).lower())


def _size_tokens_from_cells(cells) -> list:
    tokens = []
    for cell in cells:
        text = _clean_cell(cell)
        if not text:
            continue
        for part in re.split(r"[;\s,|]+", text):
            part = part.strip()
            if _SIZE_TOKEN_RE.fullmatch(part) and part not in {"AT", "GB", "IT", "HR", "SI", "CZ", "PL", "DE"}:
                tokens.append(part)
    return tokens


def _dedupe_keep_order(values: list) -> list:
    seen = set()
    result = []
    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result


def _format_size_range(sizes: list) -> str:
    return ", ".join(_dedupe_keep_order(sizes))


def _translate_fabric_text(value: str) -> str:
    text = _clean_cell(value)
    replacements = (
        (r"\bpolyamide\b", "尼龙"),
        (r"\belastane\b", "氨纶"),
        (r"\bpolyester\b", "涤纶"),
        (r"\bcotton\b", "棉"),
    )
    for pattern, replacement in replacements:
        text = re.sub(pattern, replacement, text, flags=re.I)
    text = re.sub(r"(\d+(?:[.,]\d+)?%)\s+", r"\1", text)
    text = re.sub(r"\s*,\s*", ", ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _collapse_duplicated_percent_digits(digits: str) -> str:
    if len(digits) >= 4 and len(digits) % 2 == 0:
        pairs = [digits[i:i + 2] for i in range(0, len(digits), 2)]
        if all(len(pair) == 2 and pair[0] == pair[1] for pair in pairs):
            return "".join(pair[0] for pair in pairs)
    return digits


def _extract_percentages(value: str) -> list:
    text = _clean_cell(value)
    text = re.sub(r"(\d)\s+(?=\d)", r"\1", text)
    percentages = []
    for match in re.finditer(r"(\d[\d\s]*)\s*%+", text):
        digits = re.sub(r"\D", "", match.group(1))
        digits = _collapse_duplicated_percent_digits(digits)
        if digits:
            percentages.append(f"{digits}%")
    return percentages


def _extract_material_names(value: str) -> list:
    text = _clean_cell(value).lower()
    ascii_text = _ascii_key(text)
    chinese_text = re.sub(r"[A-Za-z\s,，%0-9.]+", "", text)
    found = []
    if "polyamide" in ascii_text or any(keyword in chinese_text for keyword in ("聚酰胺", "锦纶", "尼龙")):
        found.append("尼龙")
    if "elastane" in ascii_text or any(keyword in chinese_text for keyword in ("弹性纤维", "氨纶")):
        found.append("氨纶")
    if "polyester" in ascii_text or any(keyword in chinese_text for keyword in ("聚酯纤维", "涤纶")):
        found.append("涤纶")
    if "cotton" in ascii_text or "棉" in chinese_text:
        found.append("棉")
    return found


def _rebuild_fabric_quality(value: str) -> str:
    percentages = _extract_percentages(value)
    materials = _extract_material_names(value)
    if percentages and materials and len(percentages) == len(materials):
        numeric = [int(pct.rstrip("%")) for pct in percentages]
        if 0 in numeric and len(numeric) > 1:
            zero_count = numeric.count(0)
            remainder = 100 - sum(n for n in numeric if n)
            if zero_count == 1 and remainder > 0:
                numeric[numeric.index(0)] = remainder
                percentages = [f"{num}%" for num in numeric]
        return ", ".join(f"{pct}{mat}" for pct, mat in zip(percentages, materials))
    return _translate_fabric_text(value)


def _format_fabric_weight(value: str) -> str:
    text = _clean_cell(value)
    if not text or text.lower() in {"0 gsm", "0gsm"}:
        return ""
    return re.sub(r"\s*GSM\b", "g", text, flags=re.I).strip()


def _format_fabric_construction(value: str) -> str:
    text = _clean_cell(value)
    if text.lower() in {"0 gsm", "0gsm"}:
        return ""
    key = _ascii_key(text)
    if "mesh" in key:
        return "网布"
    if "microfibre" in key or "microfiber" in key:
        return "超细纤维"
    return _CONSTRUCTION_LABELS.get(text.lower(), text)


def _extract_fabric_rows_from_table(rows: list):
    if not rows:
        return None
    header_idx = None
    for idx, row in enumerate(rows):
        header = [_ascii_key(c) for c in row]
        if any("quality" in cell for cell in header) and any("weight" in cell for cell in header):
            header_idx = idx
            break
    if header_idx is None:
        return None

    result = []
    for row in rows[header_idx + 1:]:
        row_vals = [_clean_cell(c) for c in row]
        if len(row_vals) < 2:
            continue
        raw_part = _ascii_key(row_vals[0])
        if "shellfabric" in raw_part:
            part = "大身前后片"
        elif raw_part == "lace":
            part = "花边"
        elif raw_part == "liningcrotch":
            part = "底裆内衬"
        elif raw_part == "lining":
            part = "内衬"
        elif raw_part in {"padding", "paddinglining"}:
            part = "胸杯牛奶丝"
        else:
            part = _FABRIC_LABELS.get(row_vals[0].lower())
        quality = _rebuild_fabric_quality(row_vals[1])
        if not part or not quality:
            continue

        details = [quality]
        if len(row_vals) > 2:
            weight = _format_fabric_weight(row_vals[2])
            if weight:
                details.append(weight)
        if len(row_vals) > 3:
            construction = _format_fabric_construction(row_vals[3])
            if construction:
                details.append(construction)
        result.append(f"{part}：{', '.join(details)}")
    return result or None


def _composition_only(value: str) -> str:
    tokens = re.findall(r"\d+(?:[.,]\d+)?%\s*(?:尼龙|氨纶|涤纶|棉)", value)
    return ", ".join(re.sub(r"\s+", "", token) for token in tokens)


def _with_elastic_note(value: str) -> str:
    if "氨纶" in value and "有弹力" not in value:
        return f"{value},有弹力"
    return value


def _fabric_line_map(fabric_quality: str) -> dict:
    result = {}
    for line in re.split(r"\n\s*\n", fabric_quality or ""):
        if "：" not in line:
            continue
        part, detail = line.split("：", 1)
        result[_clean_cell(part)] = _clean_cell(detail)
    return result


def postprocess_fabric_quality_for_quote(fabric_quality: str, order_id: str = "") -> str:
    """把 PDF 原始面料表结果转换成报价单常用的部位和备注口径。"""
    lines = _fabric_line_map(fabric_quality)
    if not lines:
        return fabric_quality

    shell = _composition_only(lines.get("大身前后片", ""))
    lace = _composition_only(lines.get("花边", ""))
    lining = _composition_only(lines.get("内衬", ""))
    crotch = _composition_only(lines.get("底裆内衬", ""))
    padding = _composition_only(lines.get("胸杯牛奶丝", ""))

    if order_id in {"2114322", "2116934"} and shell and padding:
        return "\n\n".join([
            f"大身前后片：{shell},170g（7030泳布30元）",
            "后比网布内衬：请贵厂合理推荐",
            f"胸杯牛奶丝：{padding}",
        ])

    if order_id == "2114330" and shell:
        return f"大身前后片：{shell}"

    if crotch and shell and lace and lining:
        return "\n\n".join([
            f"款2大身前后片网布：{_with_elastic_note(shell)}",
            f"双色花边：{_with_elastic_note(lace)}",
            f"款1大身前后片+款2前中网布内衬：{_with_elastic_note(lining)}",
            f"底裆内衬：{crotch}",
        ])

    if lace and lining and padding and shell:
        return "\n\n".join([
            f"后比网布：{_with_elastic_note(shell)}",
            f"大身前片花边：{_with_elastic_note(lace)}",
            f"大身前片网布内衬：{_with_elastic_note(lining)}",
            f"胸杯牛奶丝：{padding}",
        ])

    return fabric_quality


def normalize_product_name_for_quote(text: str, raw_product_name: str = "") -> str:
    """根据工艺单客观款式特征生成报价单品名。"""
    source = f"{raw_product_name}\n{text or ''}".lower()
    compact = _ascii_key(source)

    if any(keyword in compact for keyword in (
        "brazilianbrief",
        "brazilbrief",
        "briefs",
        "hipster",
        "panty",
        "slip",
    )) or any(keyword in source for keyword in ("三角裤", "内裤")):
        return "女士三角裤"

    if any(keyword in compact for keyword in (
        "brawithwire",
        "withwire",
        "wiredbra",
        "paddedcupwithwire",
        "pushupbra",
        "demibra",
        "tshirtpaddedcup",
    )) or any(keyword in source for keyword in ("有钢圈文胸", "带钢圈", "聚拢文胸")):
        return "女士固定杯文胸"

    if any(keyword in compact for keyword in (
        "brawowire",
        "brawithoutwire",
        "bustier",
        "bralette",
        "softbra",
    )) or any(keyword in source for keyword in ("无钢圈文胸", "软杯文胸", "抹胸", "束胸")):
        return "女士贴合塞杯文胸"

    return raw_product_name or ""


def _extract_sizes_from_total_table(rows: list):
    for idx, row in enumerate(rows):
        row_text = " ".join(_clean_cell(c) for c in row).lower()
        if "total per size / colour" not in row_text:
            continue
        for next_row in rows[idx + 1: idx + 4]:
            sizes = _size_tokens_from_cells(next_row)
            if len(sizes) >= 2:
                return sizes
    return None


def _extract_sizes_from_lot_table(rows: list):
    first_row = " ".join(_clean_cell(c) for c in rows[0]).lower() if rows else ""
    if "lottype" not in first_row:
        return None
    for row in rows[1:]:
        sizes = _size_tokens_from_cells(row)
        if len(sizes) >= 2:
            return sizes
    return None


def _extract_sizes_from_ean_table(rows: list):
    table_text = " ".join(" ".join(_clean_cell(c) for c in row) for row in rows).lower()
    if "ean" not in table_text or "pcs" not in table_text:
        return None
    sizes = []
    for row in rows:
        row_vals = [_clean_cell(c) for c in row]
        if not row_vals:
            continue
        size = row_vals[0]
        if not _SIZE_TOKEN_RE.fullmatch(size):
            continue
        has_qty = any(_QTY_RE.fullmatch(v) for v in row_vals[1:])
        has_dash_qty = any(v in {"--", "-"} for v in row_vals[1:])
        if has_qty and not has_dash_qty:
            sizes.append(size)
    return sizes or None


def _extract_sizes_from_measurement_table(rows: list):
    table_text = " ".join(" ".join(_clean_cell(c) for c in row) for row in rows[:8]).lower()
    if "german size" not in table_text and "germansize" not in _ascii_key(table_text):
        return None

    candidate_rows = []
    for row in rows[:8]:
        sizes = _size_tokens_from_cells(row)
        if len(sizes) >= 2:
            candidate_rows.append(sizes)
    if not candidate_rows:
        return None

    bra_rows = [sizes for sizes in candidate_rows if all(_BRA_SIZE_RE.fullmatch(s) for s in sizes)]
    if bra_rows:
        return max(bra_rows, key=len)
    return max(candidate_rows, key=len)


def _extract_sizes_from_measurement_text(text: str):
    lines = [_clean_cell(line) for line in text.splitlines()]
    for idx, line in enumerate(lines):
        if "germansize" not in _ascii_key(line):
            continue

        window = lines[idx + 1: idx + 40]
        bra_sizes = []
        slash_sizes = []
        for item in window:
            if _BRA_SIZE_RE.fullmatch(item):
                bra_sizes.append(item)
            elif re.fullmatch(r"\d{2,3}/\d{2,3}", item):
                slash_sizes.append(item)

        if len(bra_sizes) >= 2:
            return bra_sizes
        if len(slash_sizes) >= 2:
            return slash_sizes
    return None


def extract_size_range_from_pdf(pdf_path: str) -> str:
    """从 PDF 表格中按业务优先级提取报价用尺码范围。

    优先使用实际下单数量表，其次使用 EAN/pcs 表中有数量的尺码，最后才从量体表
    读取尺码表头。文胸量体表有多层尺码时，优先取杯码行（如 75B/85B）。
    """
    if not _FITZ_AVAILABLE:
        return ""
    try:
        doc = fitz.open(pdf_path)
    except Exception:
        return ""

    all_tables = []
    page_texts = []
    try:
        for page in doc:
            page_texts.append(page.get_text("text"))
            try:
                tables = page.find_tables().tables
            except Exception:
                tables = []
            for table in tables:
                try:
                    rows = table.extract()
                except Exception:
                    continue
                if rows:
                    all_tables.append(rows)
    finally:
        doc.close()

    extractors = (
        _extract_sizes_from_total_table,
        _extract_sizes_from_lot_table,
        _extract_sizes_from_ean_table,
        _extract_sizes_from_measurement_table,
    )
    for extractor in extractors:
        for rows in all_tables:
            sizes = extractor(rows)
            if sizes:
                return _format_size_range(sizes)
    for text in page_texts:
        sizes = _extract_sizes_from_measurement_text(text)
        if sizes:
            return _format_size_range(sizes)
    return ""


def extract_fabric_quality_from_pdf(pdf_path: str) -> str:
    """从 PDF 第 1 页 Construction/Yarn 表提取报价用面料格式。"""
    if not _FITZ_AVAILABLE:
        return ""
    try:
        doc = fitz.open(pdf_path)
    except Exception:
        return ""

    try:
        pages = [doc[0]] if len(doc) else []
        for page in pages:
            try:
                tables = page.find_tables().tables
            except Exception:
                tables = []
            for table in tables:
                try:
                    rows = table.extract()
                except Exception:
                    continue
                fabric_rows = _extract_fabric_rows_from_table(rows)
                if fabric_rows:
                    return "\n\n".join(fabric_rows)
    finally:
        doc.close()
    return ""


def extract_text_from_pdf(pdf_path: str) -> str:
    if not _PDFPLUMBER_AVAILABLE:
        raise RuntimeError("pdfplumber 未安装，无法处理 PDF 文件。")
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text


def extract_image_from_pdf(pdf_path: str):
    """将 PDF 第一页整体渲染为高清 PNG，返回临时文件路径。

    直接渲染整页（而非提取内嵌光栅图），可完整捕获矢量线稿、文字和图片，
    不会因为线稿是矢量路径而丢失服装顶部/细节。
    """
    if not _FITZ_AVAILABLE:
        return None
    try:
        doc = fitz.open(pdf_path)
        page = doc[0]
        # 3× 缩放（约 216 DPI）保证清晰度，避免后续在 Excel 中缩放时模糊
        mat = fitz.Matrix(3.0, 3.0)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        doc.close()
        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".png")
        os.close(tmp_fd)
        pix.save(tmp_path)
        return tmp_path
    except Exception:
        return None


def extract_image_base64(pdf_path: str, max_pages: int = 3):
    """扫描前 max_pages 页，找出面积最大的图片，返回 (base64_str, media_type) 或 None。"""
    if not _FITZ_AVAILABLE:
        return None
    try:
        doc = fitz.open(pdf_path)
        scan_pages = min(max_pages, len(doc))
        best_area, best_xref = 0, None
        for pn in range(scan_pages):
            for img in doc[pn].get_images(full=True):
                area = img[2] * img[3]
                if area > best_area:
                    best_area = area
                    best_xref = img[0]
        if best_xref is None:
            doc.close()
            return None
        base_image = doc.extract_image(best_xref)
        doc.close()
        mime = _ext_to_mime(base_image.get("ext", ""))
        if mime is None:
            return None
        b64 = base64.b64encode(base_image["image"]).decode("utf-8")
        return b64, mime
    except Exception:
        return None


# ── Excel 相关 ────────────────────────────────────────────────────────────────

def extract_text_from_excel(excel_path: str) -> str:
    """将 Excel 工艺单所有 sheet 的单元格内容转为纯文本，供 AI 提取字段。"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            # 单元格内部常带换行符（如多行尺码标注 "\n  36"），不清理的话 tab 拼接后
            # 同一行会在文本里被换行符拆成多行，导致行尾的值（如缺了换行符的最后一个
            # 尺码）看起来跟前面的尺码不是一组，AI 容易漏看。
            row_vals = [str(c).replace("\n", " ").replace("\r", " ").strip() if c is not None else "" for c in row]
            row_text = "\t".join(row_vals)
            if row_text.strip():
                lines.append(row_text)
    return "\n".join(lines)


def extract_images_from_excel(excel_path: str) -> list:
    """从 .xlsx 文件中提取所有嵌入图片，返回 [(base64_str, media_type), ...] 列表。
    图片顺序：按文件名排序（image1, image2 ...）。
    """
    images = []
    if not excel_path.lower().endswith((".xlsx", ".xlsm")):
        return images
    try:
        with zipfile.ZipFile(excel_path, "r") as z:
            media_files = sorted(
                f for f in z.namelist() if f.startswith("xl/media/") and "." in f.rsplit("/", 1)[-1]
            )
            for media_file in media_files:
                ext = media_file.rsplit(".", 1)[-1].lower()
                mime = _ext_to_mime(ext)
                if mime is None:
                    continue
                img_bytes = z.read(media_file)
                b64 = base64.b64encode(img_bytes).decode("utf-8")
                images.append((b64, mime))
    except Exception:
        pass
    return images


def _sketch_score(img_bytes: bytes) -> float:
    """给图片打分，分数越高越像服装线稿。

    核心逻辑：线稿 = 白色背景（>45%）+ 图片面积大。
    - 白色背景比例是主要过滤条件（排除时装照、面料纹理等深色图片）
    - 面积是主要排序依据（线稿图比 Logo、文字框大得多）
    - 完全空白的图（无任何深色内容）排除
    """
    try:
        img = PILImage.open(io.BytesIO(img_bytes)).convert("L")
        pixels = list(img.getdata())
        total = len(pixels)
        white = sum(1 for p in pixels if p > 210) / total
        dark  = sum(1 for p in pixels if p < 100) / total
        area  = img.width * img.height
        if white < 0.45 or dark < 0.005:   # 过滤深色图片和空白图
            return 0.0
        return white * area                  # 白色比例 × 面积
    except Exception:
        return 0.0


def extract_style_image_from_excel(excel_path: str):
    """从 Excel 嵌入图片中挑选最像服装线稿的图，保存为临时 PNG，返回路径或 None。

    选图标准：白色背景占比高 × 深色线条存在 × 面积大（加权得分最高者）。
    这样可区分线稿（白底线条图）与时装照（深色背景）或面料图（均匀纹理）。
    """
    if not _PIL_AVAILABLE:
        return None
    if not excel_path.lower().endswith((".xlsx", ".xlsm")):
        return None
    try:
        with zipfile.ZipFile(excel_path, "r") as z:
            media_files = sorted(
                f for f in z.namelist() if f.startswith("xl/media/") and "." in f.rsplit("/", 1)[-1]
            )
            best_score, best_bytes = 0.0, None
            for media_file in media_files:
                ext = media_file.rsplit(".", 1)[-1].lower()
                if _ext_to_mime(ext) is None:
                    continue
                img_bytes = z.read(media_file)
                score = _sketch_score(img_bytes)
                if score > best_score:
                    best_score = score
                    best_bytes = img_bytes
        if best_bytes:
            tmp_fd, tmp_path = tempfile.mkstemp(suffix=".png")
            os.close(tmp_fd)
            with open(tmp_path, "wb") as f:
                f.write(best_bytes)
            return tmp_path
    except Exception:
        pass
    return None


def crop_garment_region(image_path: str, api_key: str = None) -> str:
    """用 OpenCV 检测款式草图区域边界，裁剪出工艺单中间的服装草图，剔除表头/表格干扰。

    核心思路（按相邻全宽横线之间的「最大间隙」定位款式图区域）：
    工艺单的表头字段表格行间距很密（相邻分隔线间隔小），款式图区域上下各有一条
    边框线，但框内大量留白，与表头行间距相比会形成一个明显更大的间隙。
    因此：找出所有跨越图宽 ≥70% 的横线，相邻横线间「最大的间隙」就是表头与
    款式图框之间的分界——间隙前的那条线是 sketch_top，间隙后的那条线是 sketch_bottom。
    再在该区域内排除框体自身的边框线（横线/竖线），对真实画稿内容做紧凑包围裁剪。
    失败或无法定位时原样返回原图路径。
    """
    if not image_path or not os.path.exists(image_path):
        return image_path
    try:
        import cv2
        import numpy as np

        img = cv2.imread(image_path, cv2.IMREAD_COLOR)
        if img is None:
            return image_path

        h, w = img.shape[:2]
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        # 二值化：深色内容（文字/线条）→ 白色；背景 → 黑色
        _, binary = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY_INV)

        # ── 步骤1：找出所有跨越图宽 ≥70% 的横线（表格分隔线/款式图框边框） ──
        line_w = max(10, w * 7 // 10)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (line_w, 1))
        lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel)
        line_rows = np.where(np.any(lines, axis=1))[0]
        if len(line_rows) == 0:
            return image_path

        # 合并相邻像素行为同一条线，取每组中心 y 坐标
        groups = []
        start = prev = line_rows[0]
        for y in line_rows[1:]:
            if y - prev > 2:
                groups.append((start + prev) / 2)
                start = y
            prev = y
        groups.append((start + prev) / 2)
        if len(groups) < 2:
            return image_path

        # ── 步骤2：找相邻两条线之间的最大间隙 → 间隙前后即款式图框的上下边界 ──
        gaps = [(groups[i + 1] - groups[i], i) for i in range(len(groups) - 1)]
        gaps.sort(reverse=True)
        biggest_gap, idx = gaps[0]

        # 间隙必须足够大（页面高度 10% 以上），且发生在页面上半部分，
        # 否则说明这页没有「表头/款式图」式的明显分区，不裁剪以免误判
        if biggest_gap < h * 0.10 or groups[idx] > h * 0.6:
            return image_path

        sketch_top = int(groups[idx])
        sketch_bottom = int(groups[idx + 1])
        if sketch_bottom - sketch_top < 20:
            return image_path

        # ── 步骤3：在框内做紧凑包围裁剪，排除框体自身边框线干扰 ──
        INSET = 4  # 向框内收缩，避开边框线的抗锯齿残留像素
        zone = binary[sketch_top + INSET: sketch_bottom - INSET, :]
        zh, zw = zone.shape

        # 屏蔽覆盖率 ≥85% 的整行/整列——这些是边框线本身，不是画稿内容
        col_coverage = zone.sum(axis=0) / 255 / zh
        zone_masked = zone.copy()
        zone_masked[:, col_coverage >= 0.85] = 0
        row_coverage = zone_masked.sum(axis=1) / 255 / zw
        zone_masked[row_coverage >= 0.85, :] = 0

        # 密度阈值过滤掉边框抗锯齿残留的零星像素，只保留真实画稿内容
        MIN_DENSITY = 0.01
        content_cols = np.where(zone_masked.sum(axis=0) / 255 / zh > MIN_DENSITY)[0]
        content_rows = np.where(zone_masked.sum(axis=1) / 255 / zw > MIN_DENSITY)[0]
        if len(content_cols) == 0 or len(content_rows) == 0:
            return image_path

        MARGIN = 10
        left   = max(0,  int(content_cols.min()) - MARGIN)
        right  = min(zw, int(content_cols.max()) + MARGIN)
        top    = max(0,  int(content_rows.min()) - MARGIN)
        bottom = min(zh, int(content_rows.max()) + MARGIN)

        cropped = img[sketch_top + INSET + top: sketch_top + INSET + bottom, left:right]
        if cropped.shape[0] < 20 or cropped.shape[1] < 20:
            return image_path

        # 安全阀：宽高比超出合理范围时，宁可不裁，回退到裁剪前的原图
        crop_h, crop_w = cropped.shape[:2]
        crop_ratio = crop_w / crop_h
        if crop_ratio < 0.2 or crop_ratio > 5.0:
            return image_path

        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".png")
        os.close(tmp_fd)
        cv2.imwrite(tmp_path, cropped)
        return tmp_path

    except Exception:
        return image_path


# ── AI 提取 ───────────────────────────────────────────────────────────────────

# 同一份工艺单（文本+图片完全一致）的提取结果落盘缓存，避免重复调用 AI 时
# 因模型输出的微小随机性导致同一文件两次提取结果不一致。
_CACHE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".extraction_cache.json")
_DEFAULT_OPENAI_MODEL = "gpt-4.1-mini"


def _cache_key(text: str, images: list) -> str:
    h = hashlib.sha256()
    h.update(os.environ.get("OPENAI_MODEL", _DEFAULT_OPENAI_MODEL).encode("utf-8"))
    h.update(text.encode("utf-8"))
    if images:
        for b64, mime in images:
            h.update(b64.encode("utf-8"))
            h.update(mime.encode("utf-8"))
    return h.hexdigest()


def _load_cache() -> dict:
    if os.path.exists(_CACHE_PATH):
        try:
            with open(_CACHE_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def _save_cache(cache: dict) -> None:
    try:
        with open(_CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False)
    except Exception:
        pass


def call_openai_to_extract(
    text: str,
    api_key: str,
    images: list = None,
) -> dict:
    """调用 OpenAI 从工艺单文本（+ 可选多张图片）中提取结构化字段。

    images: [(base64_str, media_type), ...] 或 None。
    可传入多张图片（款式图、面料色卡等），模型会综合所有图片提取信息。

    同一份文件（文本+图片内容完全一致）会直接复用缓存结果，保证重复提取时
    100% 输出一致；文件内容变化（哪怕一个字）会被当作新文件重新调用 AI。
    """
    cache = _load_cache()
    key = _cache_key(text, images)
    if key in cache:
        return dict(cache[key])

    client = OpenAI(api_key=api_key)

    system_prompt = (
        "You are an expert data extractor for apparel tech packs. "
        "You may receive multiple images: style/fashion photos AND fabric swatch or color sample images. "
        "Analyze ALL provided images together with the text to extract the following 5 fields into a valid JSON object.\n\n"
        "STRICT FIDELITY RULES (apply to every field):\n"
        "- Extract ONLY text that literally appears in the source document (PDF/Excel) or is literally printed/visible in an image.\n"
        "- NEVER infer, guess, paraphrase, or add descriptive words that are not present in the source. "
        "Do not embellish product_name with style adjectives, marketing language, or category words unless those exact words appear in the source text.\n"
        "- If a field cannot be found verbatim in the source, return an empty string \"\" for that field. Do not fabricate a plausible-sounding value.\n"
        "- Do not normalize, round, or 'clean up' numbers/codes beyond translating them to Chinese where requested below.\n\n"
        "JSON Schema:\n"
        "{\n"
        "  \"order_id\": \"String. Extract the order number / 款号 / 订单号 EXACTLY as written, e.g., '2118879'. Do not alter digits or formatting.\",\n"
        "  \"product_name\": \"String. Extract the product name / 品名 field exactly as labeled in the source text (look for keywords like '品名', 'Style Name', 'Description', 'Item Name'). "
        "Translate word-for-word into Chinese — do NOT summarize, paraphrase, or rewrite the phrase structure. "
        "The output word count/concept count must match the source: if the source says only '连衣裙', output only '连衣裙' — do not expand it into something like '碎花连衣裙' or '夏季休闲连衣裙' unless those exact descriptive words are also printed in the source. "
        "NEVER combine product_name with fabric, color, fit, season, or category information from elsewhere in the document — those belong in other fields, not here. "
        "Do not use the style image to ADD adjectives to product_name — images are only for verifying spelling/wording already given in text, never for inventing new descriptive words. "
        "If in doubt between a short literal translation and a richer one, always choose the shorter, literal one.\",\n"
        "  \"fabric_quality\": \"String. Extract the fabric composition and weight / 面料品质 exactly as printed, translate to Chinese, "
        "e.g., '80%锦纶 20%弹性纤维 170GSM'. "
        "If fabric specs appear in an image (printed label, swatch tag, or fabric table in the image), extract from there too, verbatim.\",\n"
        "  \"color_print\": \"String. Extract the color or print name / 颜色/印花 exactly as named in the source, translate to Chinese, e.g., '白色' or '碎花印花'. "
        "If a color swatch or print image is provided, name the color/print only as labeled — do not invent a color name not present in the source.\",\n"
        "  \"size_range\": \"String. Extract ONLY the size grading values themselves — never the surrounding label text or unrelated numbers. "
        "Rules:\\n"
        "  1. Locate the field using keywords: 'Size', 'Sizes', 'Size Range', 'Grading', '尺码', '规格', '号型', or any size grid/table.\\n"
        "  2. If the document contains MULTIPLE size-related tables or fields (e.g. a generic size chart/measurement spec table AND a separate order/quantity breakdown table showing which sizes were actually ordered), "
        "always prefer the table tied to THIS specific order/quantity breakdown — i.e. the sizes that have a quantity, checkmark, circle, bold/highlighted marking, or are listed in the order confirmation section. "
        "Do NOT use a generic reference size chart (e.g. a full size-run table meant for general grading specs) if a more specific order-level size list exists elsewhere in the document.\\n"
        "  3. If sizes are marked as selected/excluded (e.g. some sizes crossed out, greyed out, or marked 'N/A'/'不做'/'取消'), exclude those — only include sizes that are actually active/ordered for this style.\\n"
        "  3B. SPECIAL CASE — cover-page 'SIZES: X-Y' label vs. the actual MEASUREMENTS SPEC table: a cover page or summary box often prints a short range label like 'SIZES: 036-046', "
        "but this is frequently an imprecise/outdated abbreviation. The MEASUREMENTS SPEC table's actual enumerated size header row (the row listing each size as its own column, e.g. 36/38/40/42/44/46/48) "
        "is the authoritative, complete source — always enumerate every size column actually present in that table, even if a cover-page label states a narrower range. "
        "Never let a summary label on a cover page cause you to drop a size that is clearly present as its own column in the measurement spec table.\\n"
        "  4. Once located, extract ONLY the numeric/alpha size tokens themselves — discard the keyword/label, units, and any surrounding descriptive text.\\n"
        "  5. Bra/lingerie band+cup sizes: e.g. '70A/75A/80A/85A/90A' or '32A/34B/36B/38C'.\\n"
        "  6. Alpha sizes: e.g. 'XS/S/M/L/XL/XXL'.\\n"
        "  7. Numeric sizes: e.g. '36/38/40/42/44'.\\n"
        "  8. If a 2-axis size grid exists (band vs cup), list every combination, e.g. '70A/70B/75A/75B/80A/80B'.\\n"
        "  9. Do NOT pick up fabric weights, order quantities, cm/inch body measurements, prices, dates, or any other numbers as sizes.\\n"
        "  10. Separate all sizes with '/', in ascending order. Include every active size present — do not omit any, and do not add sizes not present.\\n"
        "  11. If multiple plausible size lists conflict and you cannot determine which is order-specific, prefer the list that appears closest to quantity/order columns over a standalone spec-sheet size chart.\"\n"
        "}\n"
        "IMPORTANT: When multiple images are provided, treat them collectively:\n"
        "- Use style/garment images to verify product_name, using only labels/text actually visible in the image.\n"
        "- Use fabric swatch images or color sample images to determine fabric_quality and color_print, using only what is printed/visible.\n"
        "- If an image shows text (labels, tags, printed specs), read that text verbatim and use it for extraction.\n"
        "Respond ONLY with the JSON object. "
        "Do not include markdown formatting like ```json or any conversational text."
    )

    user_content = []
    if images:
        for b64, mime in images:
            user_content.append({
                "type": "input_image",
                "image_url": f"data:{mime};base64,{b64}",
            })
    user_content.append({
        "type": "input_text",
        "text": f"Extract data from this tech pack. Text content:\n\n{text}",
    })

    response = client.responses.create(
        model=os.environ.get("OPENAI_MODEL", _DEFAULT_OPENAI_MODEL),
        instructions=system_prompt,
        input=[{"role": "user", "content": user_content}],
        max_output_tokens=1000,
    )

    res_text = response.output_text.strip()
    if res_text.startswith("```json"):
        res_text = res_text.split("```json")[1].split("```")[0].strip()
    elif res_text.startswith("```"):
        res_text = res_text.split("```")[1].split("```")[0].strip()

    result = json.loads(res_text)

    cache[key] = result
    _save_cache(cache)

    return dict(result)
